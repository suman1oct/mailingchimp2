# python imports

# django imports
from django.core.urlresolvers import reverse_lazy
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.contrib.messages.views import SuccessMessageMixin
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.http import HttpResponseRedirect
from django.shortcuts import render, redirect
from django.template.loader import render_to_string
from django.utils import timezone
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.views import View, generic

# third-party packages
from openpyxl import load_workbook

# local imports
from chimp_users.models import UserProfile
from .forms import CreateCampaignForm, EditCampaignForm, MailingListForm
from .models import UserCampaign, MailingList, Template
from .mixins import ValidCampaignMixin, ValidMailingListMixin



class CreateCampaignView(LoginRequiredMixin, generic.FormView):
	"""
	Create Campaign view
	"""
	template_name = 'campaign/create_campaign.html'
	success_url = reverse_lazy('campaign:dashboard')
	form_class = CreateCampaignForm
	
	def get_form_kwargs(self):
		"""
		pass request to the CreateCampaignForm
		"""

		kwargs = super(CreateCampaignView, self).get_form_kwargs()
		kwargs['request'] = self.request
		return kwargs
	
	def form_valid(self, form, *args, **kwargs):
		"""
		save the campaign and assign authenticate user to campaign user		
		"""

		if form.is_valid():
			campaignn = form.save(commit=False)
			campaignn.user = self.request.user
			form.save()
			messages.success(self.request, 'Campaign created  Successfully')
			return super(CreateCampaignView, self).form_valid(form, *args, **kwargs)


class EditCampaignView(SuccessMessageMixin, LoginRequiredMixin, ValidCampaignMixin, generic.UpdateView):
	"""
	Edit Campaign View
	"""

	model = UserCampaign
	form_class = EditCampaignForm
	template_name = 'campaign/edit_campaign.html'
	success_url = reverse_lazy('campaign:show_campaign')
	success_message = 'Campaign Edited Successfully'
	
	def get_form_kwargs(self):
		"""
		pass request to EditCampaignForm
		"""

		kwargs = super(EditCampaignView, self).get_form_kwargs()
		kwargs['request'] = self.request
		return kwargs



class AddMailList(LoginRequiredMixin, SuccessMessageMixin, CreateView):
	"""
	Add MailingList View
	"""

	model = MailingList
	form_class = MailingListForm
	template_name = 'campaign/add_mail_list.html'
	success_message = 'Mail list added.'
	success_url = reverse_lazy('campaign:show_mailing_list')
	

	def form_valid(self,form, *args, **kwargs):
		"""
		save the MailingList and assign authenticate user to MailingList user
		"""

		if form.is_valid():
			mailinglist = form.save(commit=False)
			mailinglist.user = self.request.user
			mailinglist.save()
			return super(AddMailList, self).form_valid(form, *args, **kwargs)

	def get_form_kwargs(self):
		"""
		pass request to MailingListForm
		"""

		kwargs = super(AddMailList, self).get_form_kwargs()
		kwargs['request'] = self.request
		return kwargs


class SendEmailView(LoginRequiredMixin, View):
	"""
	SendEmailView send email using html template and mailinglist from excel file
	"""

	def get(self, request, *args, **kwargs):
		"""
		send email to all mailinglist's users 
		"""
		campaign_obj = UserCampaign.objects.get(id=self.kwargs['pk'])
		u = UserProfile.objects.get(user=self.request.user)
		print("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
		path = campaign_obj.mail_list.file
		template_path = campaign_obj.template.file
		campaign_name = campaign_obj.campaign_name

		wb = load_workbook(filename = path)
		sheet = wb.worksheets[0]
		row_count = sheet.max_row
		column_count = sheet.max_column
		users_name = []
		receivers = []
		for i in range(2, row_count):
			for j in range(1, column_count):
				email_o = sheet.cell(row = i, column = 2).value
				try:
					validate_email(email_o)
					receivers.append(email_o)
					users_name.append(sheet.cell(row = i, column = 1).value)
				except ValidationError as e:
					print(sheet.cell(row = i, column = 1).value + " is invalid email id\n")

		sent_mail = len(users_name)	# Number of count of mailing list
		rem_mail = u.remaining_email
		
		if( sent_mail < rem_mail):
			msg_html = render_to_string(template_path, {'username': self.request.user})
			
			from django.core import mail
			connection = mail.get_connection()
			# Manually open the connection
			connection.open()

			# Construct an email message that uses the connection
			for receiver in receivers:
				email1 = mail.EmailMessage('campaign_name',msg_html,'noreply@mailBuddy',[receiver],connection=connection,)
				email1.content_subtype = "html"
				email1.send() # Send the email
			connection.close()

			total_sent_email = u.sent_email + sent_mail
			total_remaining_email = u.remaining_email - sent_mail
			u.sent_email = total_sent_email
			u.remaining_email = total_remaining_email
			u.save()
			messages.success(self.request, 'Emails sent successfully')
			return redirect('campaign:show_campaign')
		else:
			messages.success(self.request, 'Please Contact to Admin for purchasing more emails')
			return redirect('campaign:show_campaign')


class DashboardView(LoginRequiredMixin, generic.ListView):
	
	"""
	Dashboard View
	"""

	template_name = 'campaign/dashboard.html'
	model = UserCampaign

	def get_queryset(self):
		"""
		send authenticated UserCamapign object to template 
		"""

		return UserCampaign.objects.filter(user = self.request.user)


class HomepageView(generic.ListView):
	
	"""
	Homepage View
	"""

	template_name = 'campaign/homepage.html'

	def get(self, *args, **kwargs):
		"""
		if user is already logged in then redirect to dashbard
		"""

		if self.request.user.is_authenticated():
			return redirect('campaign:dashboard')
		return super(HomepageView, self).get(*args, **kwargs)
	
	def get_queryset(self):
		"""
		send all object of template to the dashboard
		"""

		return Template.objects.all()
	


class CampaignDetailView(LoginRequiredMixin,ValidCampaignMixin, generic.DetailView):
	"""
	it will send detail of a campaign
	"""

	template_name = 'campaign/campaign_detail.html'
	model = UserCampaign
	
	def get_object(self):
		"""
		send object of campaign
		"""

		# Call the superclass
		object = super(CampaignDetailView, self).get_object()
		# Record the last accessed date
		object.last_accessed = timezone.now()
		object.save()
		# Return the object
		return object


class ShowTemplateView(LoginRequiredMixin, generic.ListView):
	"""
	ShowTemplateView send all object of Tempalte to the template
	"""
	template_name = 'campaign/show_template.html'
	model = Template
	paginate_by=3


class ShowMailingListView(LoginRequiredMixin, generic.ListView):
	"""
	ShowMailingListView send all MailingList object of authenticated user
	"""

	template_name = 'campaign/show_mailing_list.html'
	paginate_by=3

	def get_queryset(self):
		"""
		This will send only authenticated user MailingList object
		"""

		return MailingList.objects.filter(user = self.request.user)


class ShowCampaignView(LoginRequiredMixin, generic.ListView):
	"""
	ShowCampaignView send all UserCampaign object of authenticated user
	"""

	template_name = 'campaign/show_campaign.html'
	paginate_by=3

	def get_queryset(self):
		"""
		This will send only authenticated users Campaign objects
		"""

		return UserCampaign.objects.filter(user = self.request.user)


class DeleteCampaignView(LoginRequiredMixin,ValidCampaignMixin, SuccessMessageMixin,generic.DeleteView):
	"""
	DeleteCampaignView for delete the Campaign object
	"""

	model = UserCampaign
	template_name = 'campaign/delete_campaign.html'
	success_url = reverse_lazy('campaign:show_campaign')
	success_message = 'Campaign Deleted Successfully'


class DeleteMailingListView(LoginRequiredMixin, ValidMailingListMixin, SuccessMessageMixin, generic.DeleteView):
	"""
	DeleteMailingListView for delete the object of MailingList
	"""
	model = MailingList
	template_name = 'campaign/delete_mailing_list.html'
	success_url = reverse_lazy('campaign:show_mailing_list')
	success_message = 'MailingList Deleted Successfully'
	
